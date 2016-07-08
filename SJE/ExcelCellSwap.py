#! python3

#########################################################################
# ExcelCellSwap.py 														#	
# Written by Kanyon Edvall 												#
# 																		#
# This program takes an Excel spreadsheet from accounting in horizontal	#
# format and converts it to vertical format for import to IFS. 			#
#########################################################################

# Functions
def clear():
	os.system('cls')
	print(' Excel Cell Orientation Converter v1.0 '.center(50, '='))
	print('Currently in Beta... will have nice GUI soon :)')
	print()

# Matching rules - This dictionary defines what the various columns map to
mRule = {'A':'',
		 'B':'',
		 'C':'A',
		 'D':'B',
		 'E':'C',
		 'F':'D',
		 'G':None,
		 'H':None,
		 'I':'F',
		 'J':None,
		 'K':None,
		 'L':None,
		 'M':'',
		 'N':None,
		 'O':None}

# Headers - This dictionary defines what headers to user for the new Excel sheet
headers = {'A':None,
		   'B':'BUDGET_PERIOD',
		   'C':'ACCOUNT',
		   'D':'CODE_B',
		   'E':'CODE_C',
		   'F':'CODE_D',
		   'G':'CODE_E',
		   'H':'CODE_F',
		   'I':'CODE_G',
		   'J':'CODE_H',
		   'K':'CODE_I',
		   'L':'CODE_J',
		   'M':'AMOUNT',
		   'N':'QUANTITY',
		   'O':'TEXT'}

# Other settings
validExt = ('.xlsx', '.xlsm', '.xltx', '.xltm')
startRow = 2

# Import necessary modules and greet user
import openpyxl, os, re, pyperclip, sys
from openpyxl.cell import get_column_letter, column_index_from_string
clear()

# First ask user for path to Excel file
while True:
	# Get file path from user
	print('Enter path to Excel file including file extension')
	print('\tEx. C:\\Users\\intern\\data.xlsx')
	#print('Path (Press Enter to quit): ', end='')
	#filePath = input()
	print(r'Path (Press Enter to quit): C:\Users\intern\Documents\test.xlsx')
	filePath = r'C:\Users\intern\Documents\test.xlsx' # For quick testing

	# Check if Enter is pressed and quit if true
	if filePath == '':
		sys.exit()
	
	# Cleanup user input
	elif not filePath.endswith('.xlsx'):
		print('\tStandard Excel extension (.xlsx) not found, append it? (Y/N) ', end='')
		userInput = input()
		if userInput.lower().startswith('y'):
			filePath += '.xlsx'

	# Check that extension is valid and file actually exists
	if not filePath.endswith(validExt):
			clear()
			print('File format not supported. Supported formats are: .xlsx, .xlsm, .xltx, and .xltm')
			print()
	else:
		if not os.path.exists(filePath):
			clear()
			print('\tFile not found at ' + filePath + '. Please try again')
			print()
		else:			
			wb = openpyxl.load_workbook(filePath)
			wbName = os.path.basename(filePath)
			break
clear()
print('\tFile found. Loading ' + wbName + '...')
print()

# Load user specified sheet
while True:
	# Get sheet name from user
	print('Enter the name of the Excel sheet (Type List Sheets to see all available sheets)')
	print('Sheet Name (or press Enter for Active Sheet): ', end='')
	userInput = input()

	# Check input for errors or special phrases
	if userInput.lower() == 'list sheets':
		clear()
		print('Available sheets for ' + wbName + ': ', end='')
		print(', '.join(wb.get_sheet_names()))
		print()
	elif userInput == '':
		sheet = wb.active
		break
	else:
		try:
			sheet = wb.get_sheet_by_name(userInput)
			break
		except KeyError:
			clear()
			print(userInput + ' not found in ' + wbName)
			print()
clear()
print('\t' + sheet.title + ' selected.')
print()

# Prep for column swap
while True:
	print('What year should be printed to the Excel sheet?')
	print('Year: ', end='')
	year = input()

	# Check input for validity
	if not year.isdigit():
		clear()
		print('Date must be a number (Ex 2016)')
		print()
	elif len(year) != 4:
		clear()
		print('Please enter the full date (Ex 2016 not 16)')
		print()
	else:
		break
clear()
print('Using year ' + year + '.')
print()

# Ask user what they want to do
while True:
	print('Enter Swap to begin converting spreadsheet,')
	print('\t  Header to see header settings for new Excel sheet, or ')
	print('\t  Map to see column mapping: ', end='')
	userInput = input().lower()

	# Check input for validity
	if userInput == 'swap':
		clear()
		print('Preparing to convert sheet...')
		break
	elif userInput == 'header':
		clear()
		print('Current header settings: (Press Enter to exit)')
		print('\n'.join([' = '.join([key, str(val)]) for key, val in sorted(headers.items())]))
		userInput = input()
		clear()
	elif userInput == 'map':
		clear()
		print('Current column mapping table: (Press Enter to exit)')
		print('\n'.join([' -> '.join([key, str(val)]) for key, val in sorted(mRule.items())]))
		userInput = input()
		clear()
	else:
		clear()
		print(userInput + ' is not a command. Valid commands are Swap, Header, and Map')
		print()

# Swap columns based on matching rules
# Prep workbook and sheet
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet.title = 'Swapped Sheet'

# Format sheet with correct headers
for key, value in sorted(headers.items()):
	sheet.cell(row=startRow, column=column_index_from_string(key)).value = value

# Pull data from mRule defined place



# Save the new file
print('Path: ' + filePath)
print(os.path.dirname(filePath))
print(os.path.basename(filePath))
wb.save(os.path.dirname(filePath) + 'Swapped' + wbName)