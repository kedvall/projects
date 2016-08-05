#! python3
#########################################################################################
# IFS Importer.py                                                                       #
# Written by Kanyon Edvall                                                              #
#                                                                                       #
# This program allows the user to easily import data into IFS                           #
# More IFS functionality may be added later                                             #
#########################################################################################


#************************************ Program Setup ************************************#
# Import everything
import sys, os, re, subprocess, getpass, openpyxl, pyautogui, pyperclip, base64, base64ico
import tkinter.messagebox
from tkinter import *
from tkinter import ttk, filedialog
from openpyxl.cell import get_column_letter, column_index_from_string
from time import sleep

# Set Up GUI
root = Tk()
root.title('IFS Importer') # Set the name
style = ttk.Style() # Set the style


############################################################################################################################
# Class declaration
class FileSelection:
# File Selection Frame (Upper left)

	def __init__(self):
		### Frame setup ###
		fileFrame = ttk.LabelFrame(root, text='File Selection: ', padding='3 3 12 12') # Make a themed frame to hold objects
		fileFrame.grid(columnspan=6, row=0, pady=10, sticky='N W S E')

		# Required variables
		FileSelection.filePath = StringVar()
		self.selectedSheet = StringVar()
		self.fileDisp = StringVar()

		# File opening options
		self.fileOpt = options = {}
		options['defaultextension'] = '.xlsx'
		options['filetypes'] = [('Excel', '.xlsx'), ('Spreadsheet', '.xlsm'), ('Spreadsheet', '.xltx'), ('Spreadsheet', '.xltm')]
		options['initialdir'] = 'C:\\Users\\' + getpass.getuser() + '\\Documents'
		options['parent'] = fileFrame
		options['title'] = 'Select a File to Use'

		# Interface elements
		self.fileBtn = ttk.Button(fileFrame, text='Select File:', style='fileBtn.TButton', command=self.askDir)
		root.bind('<Return>', self.eventPass)
		self.fileBtn.focus_set()
		self.fileBtn.grid(columnspan=3, row=1, sticky=E)
		self.fileEntry = ttk.Entry(fileFrame, width=50, textvariable=self.filePath)
		self.fileEntry.grid(columnspan=3, column=4, row=1, sticky=W)

		ttk.Label(fileFrame, text='Select Sheet:').grid(columnspan=3, row=2, sticky=E)
		self.sheetCBox = ttk.Combobox(fileFrame, textvariable=self.selectedSheet, width=30, state='readonly')
		self.sheetCBox['values'] = ''
		self.sheetCBox.grid(columnspan=3, column=4, row=2, sticky=W)

		self.loadBtn = ttk.Button(fileFrame, text='Load File', style='loadBtn.TButton')
		self.loadBtn.state(['disabled'])
		self.loadBtn.grid(columnspan=3, row=3, sticky='W S')
		ttk.Label(fileFrame, textvariable=self.fileDisp).grid(columnspan=3, column=4, row=3, sticky='W S')

		for child in fileFrame.winfo_children(): child.grid_configure(padx=5, pady=10)


	def eventPass(self, event):
		self.askDir()


	def askDir(self):
		# Open ask directory dialog and set path
		self.filename = filedialog.askopenfilename(**self.fileOpt)
		if self.filename:
			FileSelection.filePath.set(self.filename)
			root.bind('<Return>', self.loadFile)
			self.loadBtn.bind('<Button-1>', self.loadFile)
			self.loadBtn.state(['!disabled'])
			self.loadBtn.focus_set()
			self.loadWorkbook()
		style.configure('fileBtn.TButton', relief=RAISED)


	def loadWorkbook(self):
	# Load workbook and setup sheet selection
		try:
			FileSelection.wb = openpyxl.load_workbook(FileSelection.filePath.get())
			self.sheetCBox['values'] = FileSelection.wb.get_sheet_names()
			self.sheetCBox.current(self.sheetCBox['values'].index(FileSelection.wb.active.title))
		except FileNotFoundError:
			self.fileDisp.set('Could not load file at ' + self.filename)


	def loadFile(self, event):
		# Load selected sheet from workbook
		try:
			FileSelection.sheet = FileSelection.wb.get_sheet_by_name(self.selectedSheet.get())
			self.fileDisp.set('Successfully loaded ' + str(self.selectedSheet.get()) + '.')
			ColumnSelection.mainFrame.grid()
		except KeyError:
			self.fileDisp.set('Error loading ' + str(self.selectedSheet.get()) + '!')
		

############################################################################################################################
class ColumnSelection:
# Parameter selection frame (Upper right)

	def __init__(self):
		### Frame setup ###
		ColumnSelection.mainFrame = ttk.LabelFrame(root, text='Import Options: ', padding='3 3 12 12')
		ColumnSelection.mainFrame.grid(columnspan=6, pady=10, row=3, sticky='N W S E')

		### Required variables ###
		ColumnSelection.columnsToImportDict = {} # Dictionary for user entered columns
		ColumnSelection.columnSV = StringVar() # Variable to hold entered column
		ColumnSelection.overwriteSV = StringVar() # Determines if existing text will be overwritten
		ColumnSelection.clearSV = StringVar() ### for testing ###
		ColumnSelection.skippedList = []
		ColumnSelection.mismatchList = []

		### Interface elements ###
		# Subframes to hold various elements
		ColumnSelection.titleFrame = ttk.Frame(ColumnSelection.mainFrame)
		ColumnSelection.titleFrame.pack(side=TOP, fill=X, expand=True)
		ColumnSelection.entryFrame = ttk.Frame(ColumnSelection.mainFrame)
		ColumnSelection.entryFrame.pack(fill=X, expand=True)
		ColumnSelection.buttonFrame = ttk.Frame(ColumnSelection.mainFrame)
		ColumnSelection.buttonFrame.pack(side=BOTTOM, fill=X, expand=True)

		ColumnSelection.vcmd = ColumnSelection.titleFrame.register(self.validatePNCol) # Register validate command on new frame

		# Title frame
		ttk.Label(ColumnSelection.titleFrame, text='Enter Excel column to search for IFS Inventory Part number:').grid(columnspan=4, row=0, sticky=W) # Column selection label
		ColumnSelection.columnEntry = ttk.Entry(ColumnSelection.titleFrame, width=17, textvariable=ColumnSelection.columnSV, validate='key', validatecommand=(ColumnSelection.vcmd, '%P'))
		ColumnSelection.columnEntry.grid(columnspan=2, column=4, row=0, sticky=W)
		
		ColumnSelection.divider = ttk.Label(ColumnSelection.titleFrame, text='')
		ColumnSelection.divider.grid(row=2, sticky='N S E W') # Divider

		# Instruction label
		ColumnSelection.titleLbl = ttk.Label(ColumnSelection.titleFrame, text='Select columns to import data from:')
		ColumnSelection.titleLbl.grid(row=3, sticky=W) # Instruction label

		# Add padding for title frame
		for child in ColumnSelection.titleFrame.winfo_children(): 
			child.grid_configure(padx=5, pady=5)
		ColumnSelection.titleLbl.grid_configure(pady=0)
		ColumnSelection.divider.grid_configure(pady=0)

		ttk.Label(ColumnSelection.buttonFrame, text='').pack(side=TOP) # Divider

		# Create a add button
		ColumnSelection.addButton = ttk.Button(ColumnSelection.buttonFrame, text='Add Another Import Column', command=self.addImport)
		ColumnSelection.addButton.pack(side=LEFT, anchor=W, pady=5, padx=5)
		ColumnSelection.addButton.configure(state='disabled')

		# Create Import button
		ColumnSelection.importButton = ttk.Button(ColumnSelection.buttonFrame, text='Start Import', command=self.startImport)
		ColumnSelection.importButton.pack(side=RIGHT, anchor=E, pady=5, padx=5)
		ColumnSelection.importButton.configure(state='disabled')

		# Clear check box (for testing) 
		ColumnSelection.clearCB = ttk.Checkbutton(ColumnSelection.buttonFrame, text='Clear Values', variable=ColumnSelection.clearSV, onvalue='clear', offvalue='ignore') ### for testing ###
		ColumnSelection.clearCB.pack(side=RIGHT, anchor=E, pady=5, padx=5)

		# Overwrite Existing check box
		ColumnSelection.overwriteCB = ttk.Checkbutton(ColumnSelection.buttonFrame, text='Overwrite Existing', command=self.overwriteWarning, variable=ColumnSelection.overwriteSV, onvalue='overwrite', offvalue='ignore') # Overwrite existing IFS text (if present)
		ColumnSelection.overwriteCB.pack(side=RIGHT, anchor=E, pady=5, padx=5)

		### Other Tasks ###
		ImportColumn() # Add entry field
		ColumnSelection.mainFrame.grid_remove() # Temporarily hid this frame


	def overwriteWarning(self):
		if ColumnSelection.overwriteSV.get() == 'overwrite':
			overwrite = tkinter.messagebox.askquestion('Overwrite confirmation', "This will OVERWRITE existing text in IFS!\nPlease be ABSOLUTELY SURE you know what you're doing!\nAsk someone before proceeding or deselect this option.\n\nDo you really want to proceed with overwrite on?")
			if overwrite != 'yes':
				ColumnSelection.overwriteCB.invoke()


	def addImport(self):
	# Add another import entry
		ImportColumn() # Add another button
		ColumnSelection.addButton.configure(state='disabled')


	def startImport(self):
		if ColumnSelection.columnSV.get() != '':
			WriteData()
		else:
			tkinter.messagebox.showerror('Error', 'Please enter column to search for part ID')


	def validatePNCol(self, columnValue):
		if not (columnValue.isalpha() or columnValue == ''):
			return False

		elif columnValue != '':
			try:
				column_index_from_string(columnValue.upper())
			except ValueError:
				return False

		return True


############################################################################################################################
class ImportColumn:
# Class to handle creation of import instances

	def __init__(self):
	# Create and display everything needed for user to enter import column name
		### Required Variables ###
		self.columnSV = StringVar() # Variable to hold entered column
		self.entryFieldID = '' # Variable for the name of the IFS entry field

		### Interface Setup ###
		self.entryFrame = ttk.Frame(ColumnSelection.mainFrame, padding='3 3') # Frame for easy layout
		self.entryFrame.pack(anchor=W)

		self.vcmd = self.entryFrame.register(self.validateColumnEntry) # Register validate command on new frame

		ttk.Label(self.entryFrame, text='Enter column to import data from:').grid(columnspan=4, row=0) # Column selection label

		# Column selection entry
		self.columnEntry = ttk.Entry(self.entryFrame, width=17, textvariable=self.columnSV, validate='all', validatecommand=(self.vcmd, '%V', '%P'))
		self.columnEntry.grid(columnspan=2, column=4, row=0)
		self.setPlaceholder() # Set initial placeholder text

		# IFS entry field selection button
		self.fieldSelectBtn = ttk.Button(self.entryFrame, text='Select IFS Entry Field', command=self.setField)
		self.fieldSelectBtn.configure(state='disabled')
		self.fieldSelectBtn.grid(column=6, row=0)

		# Add spacing to all widgets within this frame
		for child in self.entryFrame.winfo_children(): 
			child.grid_configure(padx=4, pady=5)


	def removeImportEntry(self):
	# Remove current import entry
		del self.columnSV
		self.entryFrame.destroy()


	def setField(self):
		# Launch program to get IFS entry field ID
		self.entryFieldID = FieldSelection.getField(FieldSelection)

		if (self.entryFieldID):
			# Change Field Select button text
			self.fieldSelectBtn.configure(text='Change Selection')

		# Update necessary variables
		self.updateVars()

		# Enable search
		ColumnSelection.importButton.configure(state='enabled')


	def updateVars(self):
		if (self.columnSV.get() != 'Column: A to XFD' and self.columnSV.get() != ''):
			ColumnSelection.columnsToImportDict[self.columnSV.get()] = [] # Add column name to dict with list as key
			ColumnSelection.columnsToImportDict[self.columnSV.get()].append(self) # Append instance to dict
			ColumnSelection.columnsToImportDict[self.columnSV.get()].append(self.entryFieldID) # Add selected IFS field ID to dict


	def validateColumnEntry(self, reason, columnValue):
	# Called on entry state change, allows or denies edit. Return True to allow edit, False to disallow
		if reason == 'focusin':
			self.remPlaceholder()
			return True

		elif reason == 'focusout':
			self.setPlaceholder()
			return True

		elif reason == 'key':
			if not (columnValue.isalpha() or columnValue == ''):
				return False

			elif columnValue != '':
				try:
					column_index_from_string(columnValue.upper())
				except ValueError:
					return False

			ColumnSelection.addButton.configure(state='enabled')
			self.fieldSelectBtn.configure(state='enabled')
			return True


	def setPlaceholder(self):
	# Sets the placeholder text of the entry
		if self.columnSV.get() == '': # Check if Entry is empty before setting value
			self.columnSV.set('Column: A to XFD')
			self.columnEntry.configure(foreground='grey')


	def remPlaceholder(self):
	# Removes placeholder text of the entry
		if self.columnSV.get() == 'Column: A to XFD':
			self.columnSV.set('')
			self.columnEntry.configure(foreground='black')


class FieldSelection():
# Class to handle field selection in IFS

	def getField(self):
		# Activate IFS
		try: 
			subprocess.call(['helper\ActivateIFS.exe'])
		except FileNotFoundError:
			print('Could not locate ActivateIFS.exe')

		# Clear the clipboard
		pyperclip.copy('')

		# Get Field ID
		try: 
			subprocess.call(['helper\GetControlID.exe'])
		except FileNotFoundError:
			print('Could not locate GetControlID.exe')

		# Get the entry field info
		self.FieldID = pyperclip.paste()

		# Activate IFS Importer
		try: 
			subprocess.call(['helper\ActivateImporter.exe'])
		except FileNotFoundError:
			print('Could not locate ActivateImporter.exe')

		# return the field's ID
		return self.FieldID


############################################################################################################################
class WriteData():
# Copies data from Excel column and paste it into IFS using helper scripts

	def __init__(self):
		# Make IFS window active
		self.activateWindow()

		# Iterate though spreadsheet row by row
		for rowNum in range(1, FileSelection.sheet.max_row + 1):
			validID = self.searchByID(rowNum)
			if validID:
				self.pasteData(rowNum, self.IDTag)
			else:
				pass

		# Message display on import completion
		self.completionHandle()


	def activateWindow(self):
		# Activate IFS
		try: 
			subprocess.call(['helper\ActivateInventoryPart.exe'])
		except FileNotFoundError:
			print('Could not locate ActivateInventoryPart.exe')


	def searchByID(self, rowNum):
		# Make cell object
		curCell = FileSelection.sheet.cell(row=rowNum, column=column_index_from_string(ColumnSelection.columnSV.get().upper()))

		# Ensure there is an ID associated with the current row
		if curCell.value == None:
			# print('Cell at ' + str(curCell.column) + str(curCell.row) + ' is empty, skipping') ### Debugging ###
			return False
		else:
			pyperclip.copy(str(curCell.value))
			self.IDTag = pyperclip.paste()

		# Perform a search
		pyautogui.hotkey('f3')
		sleep(0.5)
		pyautogui.hotkey('ctrl', 'v')
		pyautogui.hotkey('enter')
		sleep(0.5)

		return True

	def pasteData(self, rowNum, IDTag):
		# Iterate though all selected data entry columns
		for columnName, propertyDict in ColumnSelection.columnsToImportDict.items():
			# Check if IFS control is empty
			pyperclip.copy(propertyDict[1])
			try: 
				subprocess.call(['helper\GetControlValue.exe'])
			except FileNotFoundError:
				print('Could not locate GetControlValue.exe')
			# Store text in IFS control	
			controlText = pyperclip.paste()

			# Set focus to correct IFS control
			pyperclip.copy(propertyDict[1])
			try: 
				subprocess.call(['helper\FocusControl.exe'])
			except FileNotFoundError:
				print('Could not locate FocusControl.exe')

			# Get value of current cell
			curCell = FileSelection.sheet.cell(row=rowNum, column=column_index_from_string(columnName))
			if curCell.value != None:
				pyperclip.copy(str(curCell.value))
			else:
				pyperclip.copy('')

			### for testing ###
			if ColumnSelection.clearSV.get() == 'clear':
				pyperclip.copy('')
				pyautogui.hotkey('ctrl', 'v')
				pyautogui.hotkey('ctrl', 's')
				sleep(0.5)
				continue

			# Nothing in field, free to paste
			if controlText == '' or ColumnSelection.overwriteSV.get() == 'overwrite':
				# Paste value into IFS
				pyautogui.hotkey('ctrl', 'v')
				pyautogui.hotkey('ctrl', 's')
				sleep(0.5)

			# Data is already in there skip
			else:
				if IDTag not in ColumnSelection.skippedList:
					ColumnSelection.skippedList.append(IDTag)
				# print('Already has text, skipping') ### Debugging ###


	def completionHandle(self):
		if len(ColumnSelection.skippedList) > 0:
			self.message = 'All fields were successfully imported.\nSome values have been skipped due to preexisting text, they are listed below.\n\nSkipped IFS Parts: ' + ', '.join(ColumnSelection.skippedList)
		else:
			self.message = 'All fields were successfully imported.'

		tkinter.messagebox.showinfo('Import Results', self.message)

		self.ans = tkinter.messagebox.askquestion('Data Validation', 'Start data validation check?')
		if self.ans == 'yes':
			ValidateData()
			return
		else:
			return


class ValidateData():
# Check existing IFS value against Excel value

	def __init__(self):
		# Make IFS window active
		self.activateWindow()

		# Iterate though spreadsheet row by row
		for rowNum in range(1, FileSelection.sheet.max_row + 1):
			validID = self.searchByID(rowNum)
			if validID:
				self.checkData(rowNum, self.IDTag)
			else:
				pass

		# Message display on import completion
		self.completionHandle()


	def activateWindow(self):
		# Activate IFS
		try: 
			subprocess.call(['helper\ActivateInventoryPart.exe'])
		except FileNotFoundError:
			print('Could not locate ActivateInventoryPart.exe')


	def searchByID(self, rowNum):
		# Make cell object
		curCell = FileSelection.sheet.cell(row=rowNum, column=column_index_from_string(ColumnSelection.columnSV.get().upper()))

		# Ensure there is an ID associated with the current row
		if curCell.value == None:
			# print('Cell at ' + str(curCell.column) + str(curCell.row) + ' is empty, skipping') ### Debugging ###
			return False
		else:
			pyperclip.copy(str(curCell.value))
			self.IDTag = pyperclip.paste()

		# Perform a search
		pyautogui.hotkey('f3')
		sleep(0.5)
		pyautogui.hotkey('ctrl', 'v')
		pyautogui.hotkey('enter')
		sleep(0.5)

		return True

	def checkData(self, rowNum, IDTag):
		# Iterate though all selected data entry columns
		for columnName, propertyDict in ColumnSelection.columnsToImportDict.items():
			# Get value of IFS control
			pyperclip.copy(propertyDict[1])
			try: 
				subprocess.call(['helper\GetControlValue.exe'])
			except FileNotFoundError:
				print('Could not locate GetControlValue.exe')
			# Store text in IFS control	
			controlText = str(pyperclip.paste())

			# Get value of current cell
			curCell = FileSelection.sheet.cell(row=rowNum, column=column_index_from_string(columnName))
			excelText = str(curCell.value)

			# Compare values
			if controlText == excelText:
				# print(IDTag + ' checked. PASS' ### Debugging ###
				pass

			# Data does not match
			else:
				if IDTag not in ColumnSelection.mismatchList:
					ColumnSelection.mismatchList.append(IDTag)
				# print('Excel IFS mismatch') ### Debugging ###


	def completionHandle(self):
		if len(ColumnSelection.mismatchList) > 0:
			self.message = 'Mismatched IFS and Excel fields detected,\nthis may be due to preexisting text.\nMismatched fields are listed below.\n\nMismatched IFS IDs: ' + ', '.join(ColumnSelection.mismatchList)
		else:
			self.message = 'All fields were successfully verified.'

		tkinter.messagebox.showinfo('Import Results', self.message)

		return


############################################################################################################################
class Base64IconGen():
# Takes icon from base64 format and creates window icon
	def __init__(self, window):
		icondata = base64.b64decode(base64ico.importerIcon)
		# The temp file is icon.ico
		os.chdir('C:\\Users' + os.sep + getpass.getuser() + os.sep + 'AppData\\Local\\Temp')
		tempFile= "icon.ico"
		iconfile= open(tempFile,"wb")
		# Extract the icon
		iconfile.write(icondata)
		iconfile.close()
		window.wm_iconbitmap(tempFile)
		# Delete the tempfile
		os.remove(tempFile)


#************************************ Program Start ************************************#
### Create objects ###
Base64IconGen(root)
filePane = FileSelection()
selectionPane = ColumnSelection()

### Run GUI ###
root.mainloop()