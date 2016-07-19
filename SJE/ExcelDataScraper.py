#! python3
#########################################################################################
# ExcelDataScraper.py 																	#
# Written by Kanyon Edvall																#						
#									 													#
# This program allows you to traverse any excel sheet and find data of interest			#
# Runs windowless with a GUI made in tkinter 										   	#
#########################################################################################

#																						#
# ************************************************************************************* #
# TODO: 																				#
# - Integrate search backend into GUI 													#
# - Regex generation 																	#
# ************************************************************************************* #
#																						#

#************************************ Program Setup ************************************#
# Import Everything
import sys, os, re, openpyxl, getpass
import tkinter.messagebox
from tkinter import *
from tkinter import ttk, filedialog
from openpyxl.cell import get_column_letter, column_index_from_string


#Global Variables
offset=''
generatedPerms = []
permsFound = []
permsToSearch = []


# Set up GUI
root = Tk() # Create blank window
root.title('Excel Data Scraper') # Set the name
style = ttk.Style()


# Class declaration
class FileSelection:
# File Selection Frame (Upper left)

	def __init__(self):
		# Frame setup
		fileFrame = ttk.LabelFrame(root, text='File Selection: ', padding='3 3 12 12') # Make a themed frame to hold objects
		fileFrame.grid(columnspan=6, row=0, pady=10, sticky='N W S E')

		# Required variables
		self.filePath = StringVar()
		self.selectedSheet = StringVar()
		self.fileDisp = StringVar()

		# File opening options
		self.fileOpt = options = {}
		options['defaultextension'] = '.xlsx'
		options['filetypes'] = [('Excel', '.xlsx'), ('Spreadsheet', '.xlsm'), ('Spreadsheet', '.xltx'), ('Spreadsheet', '.xltm')]
		options['initialdir'] = 'C:\\Users\\' + getpass.getuser() + '\\Documents'
		options['parent'] = fileFrame
		options['title'] = 'Select a File to Use'

		# Interface elements
		self.fileBtn = ttk.Button(fileFrame, text='Select File:', style='fileBtn.TButton', command=self.askDir)
		root.bind('<Return>', self.eventPass)
		self.fileBtn.focus_set()
		self.fileBtn.grid(columnspan=3, row=1, sticky=E)
		self.fileEntry = ttk.Entry(fileFrame, width=50, textvariable=self.filePath)
		self.fileEntry.grid(columnspan=3, column=4, row=1, sticky=W)

		ttk.Label(fileFrame, text='Select Sheet:').grid(columnspan=3, row=2, sticky=E)
		self.sheetCBox = ttk.Combobox(fileFrame, textvariable=self.selectedSheet, width=30, state='readonly')
		self.sheetCBox['values'] = ''
		self.sheetCBox.grid(columnspan=3, column=4, row=2, sticky=W)

		self.loadBtn = ttk.Button(fileFrame, text='Load File', style='loadBtn.TButton')
		self.loadBtn.state(['disabled'])
		self.loadBtn.grid(columnspan=3, row=3, sticky='W S')
		ttk.Label(fileFrame, textvariable=self.fileDisp).grid(columnspan=3, column=4, row=3, sticky='W S')

		for child in fileFrame.winfo_children(): child.grid_configure(padx=5, pady=10)


	def eventPass(self, event):

		self.askDir()


	def askDir(self):
		# Open ask directory dialog and set path
		self.filename = filedialog.askopenfilename(**self.fileOpt)
		if self.filename:
			self.filePath.set(self.filename)
			root.bind('<Return>', self.loadFile)
			self.loadBtn.bind('<Button-1>', self.loadFile)
			self.loadBtn.state(['!disabled'])
			self.loadBtn.focus_set()
			self.loadWorkbook()
		style.configure('fileBtn.TButton', relief=RAISED)


	def loadWorkbook(self):
	# Load workbook and setup sheet selection
		try:
			FileSelection.wb = openpyxl.load_workbook(self.filePath.get())
			self.sheetCBox['values'] = FileSelection.wb.get_sheet_names()
			self.sheetCBox.current(self.sheetCBox['values'].index(FileSelection.wb.active.title))
		except FileNotFoundError:
			self.fileDisp.set('Could not load file at ' + self.filename)


	def loadFile(self, event):
		# Load selected sheet from workbook
		try:
			FileSelection.sheet = FileSelection.wb.get_sheet_by_name(self.selectedSheet.get())
			self.fileDisp.set('Successfully loaded ' + str(self.selectedSheet.get()) + '.')
		except KeyError:
			self.fileDisp.set('Error loading ' + str(self.selectedSheet.get()) + '!')
		

class SearchSelection:
# Search Selection Frame (Lower left)
	def __init__(self):
		# Frame setup
		sModeFrame = ttk.LabelFrame(root, text='Search Type: ', padding='3 3 12 12')
		sModeFrame.grid(columnspan=6, row=2, pady=10, sticky='N S W E')
		
		# Required variables
		self.searchMode = StringVar()
		self.searchMode.set('keyword')
		self.matchMode = StringVar()
		self.matchMode.set('column')
		self.radioSet()

		# Interface elements
		ttk.Label(sModeFrame, text='Select Search Mode:').grid(columnspan=6, row=1, sticky=W)
		self.keywordRBtn = ttk.Radiobutton(sModeFrame, text='Keyword', variable=self.searchMode, value='keyword', command=self.radioSet)
		self.keywordRBtn.grid(columnspan=3, row=2, sticky=W)
		self.exactRBtn = ttk.Radiobutton(sModeFrame, text='Exact Match', variable=self.searchMode, value='exact', command=self.radioSet)
		self.exactRBtn.grid(columnspan=3, column=4, row=2, sticky=W)

		ttk.Label(sModeFrame, text='').grid(column=0, row=3, sticky=(W, E)) # Divider
		ttk.Label(sModeFrame, text='Select Match Mode:').grid(columnspan=6, row=4, sticky=W)
		self.columnRBtn = ttk.Radiobutton(sModeFrame, text='From Another Column', variable=self.matchMode, value='column', command=self.radioSet)
		self.columnRBtn.grid(columnspan=3, row=5, sticky=W)
		self.offsetRBtn = ttk.Radiobutton(sModeFrame, text='Keyword Offset', variable=self.matchMode, value='offset', command=self.radioSet)
		self.offsetRBtn.grid(columnspan=3, column=4, row=5, sticky=W)

		for child in sModeFrame.winfo_children(): child.grid_configure(padx=5, pady=0)
		

	def radioSet(self):
		search = self.searchMode.get()
		match = self.matchMode.get()
		print(search + match)
	

class ParamSelection:
# Parameter selection frame (Upper right)
	def __init__(self):
		# Frame setup
		paramFrame = ttk.LabelFrame(root, text='Search Options: ', padding='3 3 12 12')
		paramFrame.grid(columnspan=6, column=7, pady=10, row=0, sticky='N W S E')

		# Required variables
		ParamSelection.searchCol = StringVar() # Holds name of column to be searched
		ParamSelection.pasteCol = StringVar() # Holds name of column to paste data into
		ParamSelection.offsetMode = StringVar() # Currently select offset mode (Radio button)
		self.offsetPtrnLbl = StringVar() # Holds text of label above pattern entry
		self.offsetPattern = StringVar() # Holds text from pattern entry field
		vcmd = paramFrame.register(self.updateHandler) # Validation binding
		self.instructionDict = {'column':'Column: A to XFD', 'char':'Must be a number (Ex 10)'}

		# Set defaults
		ParamSelection.searchCol.set('Column: A to XFD')
		ParamSelection.pasteCol.set('Column: A to XFD')
		ParamSelection.offsetMode.set('pattern')
		self.offsetPtrnLbl.set('Enter Pattern:')

		# Interface elements
		ttk.Label(paramFrame, text='Which column would you like to search?').grid(columnspan=5, row=0, sticky=E)
		self.sColEntry = ttk.Entry(paramFrame, width=17, textvariable=ParamSelection.searchCol, foreground='grey', validate='all', validatecommand=(vcmd, '%V', '%W', '%P'))
		self.sColEntry.grid(columnspan=2, column=5, row=0, sticky=W)
		
		ttk.Label(paramFrame, text='Which column would you like to copy the selected data to?').grid(columnspan=5, row=1, sticky=E)
		self.pColEntry = ttk.Entry(paramFrame, width=17, textvariable=ParamSelection.pasteCol, foreground='grey', validate='all', validatecommand=(vcmd, '%V', '%W', '%P'))
		self.pColEntry.grid(columnspan=2, column=5, row=1, sticky=W)

		ttk.Label(paramFrame, text='').grid(columnspan=7, row=2, sticky=(W, E)) # Divider
		ttk.Label(paramFrame, text='Select Offset Type:').grid(columnspan=3, row=3, sticky=W)
		ttk.Label(paramFrame, textvariable=self.offsetPtrnLbl).grid(columnspan=5, column=2, row=3, sticky=W)

		self.patternRBtn = ttk.Radiobutton(paramFrame, text='Pattern', variable=ParamSelection.offsetMode, value='pattern', command=self.radioSet)
		self.patternRBtn.grid(column=0, row=4, sticky=W)
		self.charRBtn = ttk.Radiobutton(paramFrame, text='Character Count', variable=ParamSelection.offsetMode, value='char', command=self.radioSet)
		self.charRBtn.grid(column=1, row=4, sticky=W)
		self.ptrnEntry = ttk.Entry(paramFrame, width=30, textvariable=self.offsetPattern, validate='all', validatecommand=(vcmd, '%V', '%W', '%P'))
		self.ptrnEntry.grid(columnspan=5, column=2, row=4, sticky=W)
		
		self.nameDict = {str(self.sColEntry):{'textvar':ParamSelection.searchCol, 'placeholder':'Column: A to XFD', 'entryName':self.sColEntry, 'type':'column'},
						 str(self.pColEntry):{'textvar':ParamSelection.pasteCol, 'placeholder':'Column: A to XFD', 'entryName':self.pColEntry, 'type':'column'},
						 str(self.ptrnEntry):{'textvar':self.offsetPattern, 'placeholder':'Must be a number (Ex 10)', 'entryName':self.ptrnEntry, 'type':'pattern'},
						 'radioTriggerMapping':{'textvar':self.offsetPattern, 'placeholder':'Must be a number (Ex 10)', 'entryName':self.ptrnEntry, 'type':'pattern'}}

		for child in paramFrame.winfo_children(): child.grid_configure(padx=5, pady=5)


	def radioSet(self):
		if ParamSelection.offsetMode.get() == 'pattern':
			self.offsetPtrnLbl.set('Enter Pattern:')
			self.updateHandler('radioChange', 'radioTriggerMapping', self.offsetPattern.get())
		else:
			self.offsetPtrnLbl.set('Enter Offset (# of characters):')
			self.updateHandler('radioChange', 'radioTriggerMapping', self.offsetPattern.get())


	def updateHandler(self, reason, varName, entryValue): 
	# Called on entry state change, decides where to pass task. Return True to allow edit, False to disallow
		if reason == 'radioChange': # Radio button was clicked, check new position
			if ParamSelection.offsetMode.get() == 'char':
				if (not self.validateEntry(varName, entryValue) or self.offsetPattern.get() == ''):
					self.setPlaceholder(varName, True)
			else:
				self.remPlaceholder(varName)

		elif reason == 'focusin':
			self.remPlaceholder(varName)

		elif reason == 'focusout':
			self.setPlaceholder(varName, False)

		elif reason == 'key':
			if not self.validateEntry(varName, entryValue):
				return False

		return True		


	def validateEntry(self, varName, curEntryVal):
	# Validates the entry based on entry type. Returns True if pass, False if fail
		if self.nameDict[varName]['type'] == 'column':
			if not (curEntryVal.isalpha() or curEntryVal == ''):
				return False

		elif ParamSelection.offsetMode.get() == 'char':
			if not (curEntryVal.isdigit() or curEntryVal == ''):
				return False

		return True


	def setPlaceholder(self, varName, forceSet):
	# Sets the placeholder text of the entry. Can be forced to override current text
		textvar = self.nameDict[varName]['textvar']

		if forceSet: # If force flag is set, override current value
			textvar.set(self.nameDict[varName]['placeholder'])
			self.nameDict[varName][str('entryName')].configure(foreground='grey')

		elif textvar.get() == '': # Check if Entry is empty before setting value
			if (self.nameDict[varName]['type'] == 'pattern' and ParamSelection.offsetMode.get() == 'char'): # If pattern entry, make sure char is selected
				textvar.set(self.nameDict[varName]['placeholder'])
				self.nameDict[varName][str('entryName')].configure(foreground='grey')

			elif self.nameDict[varName]['type'] == 'column': # Column entry
				textvar.set(self.nameDict[varName]['placeholder'])
				self.nameDict[varName][str('entryName')].configure(foreground='grey')


	def remPlaceholder(self, varName):
	# Removes the placeholder text of the entry
		textvar = self.nameDict[varName]['textvar']

		if textvar.get() == self.nameDict[varName]['placeholder']:
			textvar.set('')
			self.nameDict[varName][str('entryName')].configure(foreground='black')


class Search:
# Search Frame (Lower right)
	def __init__(self):
		# Outer Frame setup
		searchFrame = ttk.LabelFrame(root, text='Search: ', padding='3 3')
		searchFrame.grid(columnspan=6, column=7, row=2, pady=10, sticky='N W S E')

		# Required variables
		self.searchTerm = StringVar()
		self.selectStateText = StringVar()
		self.selectStateText.set('Unselect All')
		self.cbVals={}
		self.resultCB={}
		self.lbl={}
		
		# Interface elements
		ttk.Label(searchFrame, text='Enter search term:').grid(columnspan=2, row=2, sticky='W')
		self.termEntry = ttk.Entry(searchFrame, width=53, textvariable=self.searchTerm)
		self.termEntry.grid(columnspan=5, column=2, row=2, sticky=W)
		
		ttk.Label(searchFrame, text="A period will allow any non alphanumeric character to be substituted in it's place.").grid(columnspan=7, row=3, sticky='W')

		specialLbl = ttk.Label(searchFrame, text='Available Permutations:').grid(columnspan=3, row=4, sticky='W S')
		self.permutBtn = ttk.Button(searchFrame, text='Search Permutations', style='permutBtn.TButton')
		self.permutBtn.bind('<Button-1>', self.searchPerms)
		self.permutBtn.grid(columnspan=2, column=5, row=4, sticky=E)

		self.startSearchBtn = ttk.Button(searchFrame, text='Start Search', style='startSearchBtn.TButton')
		self.startSearchBtn.bind('<Button-1>', self.doSomething)
		self.startSearchBtn.grid(row=9, sticky=W)
		self.selectBtn = ttk.Button(searchFrame, textvariable=self.selectStateText, style='selectBtn.TButton')
		self.selectBtn.bind('<Button-1>', self.switchState)
		self.selectBtn.grid(columnspan=5, column=1, row=9)
		self.exportBtn = ttk.Button(searchFrame, text='Export Sheet', style='exportBtn.TButton')
		self.exportBtn.bind('<Button-1>', self.doSomething)
		self.exportBtn.grid(columnspan=2, column=6, row=9, sticky=E)

		for child in searchFrame.winfo_children(): child.grid_configure(padx=5, pady=10)


		# Inner Frame Setup
		self.resultFrame = ttk.Frame(searchFrame, borderwidth=5, relief='groove', padding='3 3 120 120')
		self.resultFrame.grid(columnspan=7, rowspan=3, row=5, padx=5, sticky='N W S E')

		# Required Variables
		self.results = StringVar()
		self.results.set('Waiting for search term...')

		# Interface elements
		ttk.Label(self.resultFrame, textvariable=self.results).pack(side=LEFT)


	def doSomething(self, event):
		print('Btn clicked')


	def searchPerms(self, event):
		# Setup result headings
		self.results.set('')
		self.resultFrame.configure(padding='0 0 0 0')
		self.resultHLbl = ttk.Label(self.resultFrame, text='Result:\t')
		self.resultHLbl.pack(side=LEFT)
		self.posHLbl = ttk.Label(self.resultFrame, text='Row/Col:\t')
		self.posHLbl.pack(side=LEFT)
		self.contextHLbl = ttk.Label(self.resultFrame, text='Context:')
		self.contextHLbl.pack(side=LEFT)

		# Set up frame
		self.scrollbar = Scrollbar(self.resultFrame, orient=VERTICAL)
		self.resultbox = Listbox(self.resultFrame, selectmode=MULTIPLE, yscrollcommand=self.scrollbar.set)
		self.scrollbar.config(command=self.resultbox.yview)
		self.scrollbar.pack(side=BOTTOM, fill=Y, padx=0, pady=0)
		self.resultbox.pack(side=BOTTOM, fill=BOTH, expand=1, padx=0, pady=0)

		self.resultbox.insert(END, "a list entry")

		for item in ['one', 'two', 'three', 'four']:
			self.resultbox.insert(END, item)

		for number in range(1, 125):
			self.resultbox.insert(END, str(number))

		
		# Generate permutations based on search terms
		if self.searchTerm.get() != '':
			try:
				RegexGeneration.genPerms(RegexGeneration, str(self.searchTerm.get()))
				ExcelHandler.findPerms(ExcelHandler, FileSelection, ParamSelection)
			except AttributeError:
				print('Error!')

		
		# Print results
		for i in range(len(permsFound)):
			self.cbVals[i] = StringVar()
			self.cbVals[i].set(1)
			self.resultCB[i] = ttk.Checkbutton(self.resultFrame, text=permsFound[i], variable=self.cbVals[i])
			self.resultCB[i].bind('<Button-1>', self.updatePerms)
			self.resultCB[i].grid(columnspan=2, column=0, row=[i+2], sticky=W)

		for child in self.resultFrame.winfo_children(): child.grid_configure(padx=5)


	def updatePerms(self, event):

		print('Item state changed')


	def switchState(self, event):

		if self.selectStateText.get() == 'Unselect All':
			for box in self.resultCB:
				self.cbVals[box].set(0)
			self.updatePerms(None)
			self.selectStateText.set('Select All')
		else:
			for box in self.resultCB:
				self.cbVals[box].set(1)
			self.updatePerms(None)
			self.selectStateText.set('Unselect All')


class ExcelHandler(FileSelection, ParamSelection):
# Class to handle traversing Excel sheets
	def findPerms(self, files, params):
		self.wb = files.wb
		self.sheet = files.sheet
		print(str(params.searchCol.get()).upper())
		self.searchCol = column_index_from_string(str(params.searchCol.get()).upper())

		for rowNum in range (1, self.sheet.max_row + 1):
			curCell = self.sheet.cell(row=rowNum, column=self.searchCol)
			print('Searching row: ' + str(rowNum) + ' val: ' + str(curCell.value))
			for result in RegexGeneration.permutRegex.findall(str(curCell.value)):
				if result[0] not in (' '.join(permsFound)):
					permsFound.append(result[0])


class RegexGeneration:
# Class to handle all regular expression and pattern generation
	def genPerms(self, originTerm):
	# Generate permutations to search for based on search term
		if '.' not in originTerm:
			searchStrings = originTerm.split()
			searchStrings = '([\\-_ /])*'.join(searchStrings)
			print(str(searchStrings))
			self.permutRegex = re.compile(r'(' + searchStrings + ')+', re.I)
			print(self.permutRegex.pattern)

		else:
			searchStrings = originTerm.split()
			print('1'+str(searchStrings))
			searchStrings = '([\\-_ /])*'.join(searchStrings)
			print('2'+str(searchStrings))
			searchStrings = originTerm.split('.')
			print('3'+str(searchStrings))
			searchStrings = '([\\-_ /])*'.join(searchStrings)
			print('4'+str(searchStrings))
			self.permutRegex = re.compile(r'(' + searchStrings + ')+', re.I)
			print(self.permutRegex.pattern)


#************************************ Program Start ************************************#
# Create objects
filePane = FileSelection()
sModePane = SearchSelection()
paramPane = ParamSelection()
searchPane = Search()

# Run GUI
root.mainloop()