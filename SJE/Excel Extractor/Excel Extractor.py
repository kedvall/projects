#! python3
#########################################################################################
# Excel Extractor.py                                                                    #
# Written by Kanyon Edvall                                                              #
#                                                                                       #
# This program allows the user to traverse any spreadsheet and find data of interest    #
# Runs windowless with a GUI made in tkinter                                            #
#########################################################################################


#************************************ Program Setup ************************************#
# Import Everything
import sys, os, re, inspect, openpyxl, getpass, base64, base64ico
import tkinter.messagebox
from tkinter import *
from tkinter import ttk, filedialog
from inspect import currentframe
from openpyxl.cell import get_column_letter, column_index_from_string


# Global Variables
offset=''
permsFound = []
permsToSearch = []


# Set up GUI
root = Tk() # Create blank window
root.title('Excel Extractor') # Set the name
style = ttk.Style() # Set the style


# Class declaration
############################################################################################################################
class FileSelection:
# File Selection Frame (Upper left)

	def __init__(self):
		### Frame setup ###
		fileFrame = ttk.LabelFrame(root, text='File Selection: ', padding='3 3 12 12') # Make a themed frame to hold objects
		fileFrame.grid(columnspan=6, row=0, pady=10, sticky='N W S E')

		# Required variables
		FileSelection.filePath = StringVar()
		self.selectedSheet = StringVar()
		self.fileDisp = StringVar()

		# File opening options
		self.fileOpt = options = {}
		options['defaultextension'] = '.xlsx'
		options['filetypes'] = [('Excel', '.xlsx'), ('Spreadsheet', '.xlsm'), ('Spreadsheet', '.xltx'), ('Spreadsheet', '.xltm')]
		options['initialdir'] = 'C:\\Users\\' + getpass.getuser() + '\\Documents'
		options['parent'] = fileFrame
		options['title'] = 'Select a File to Use'

		# Interface elements
		self.fileBtn = ttk.Button(fileFrame, text='Select File:', style='fileBtn.TButton', command=self.askDir)
		root.bind('<Return>', self.eventPass)
		self.fileBtn.focus_set()
		self.fileBtn.grid(columnspan=3, row=1, sticky=E)
		self.fileEntry = ttk.Entry(fileFrame, width=50, textvariable=self.filePath)
		self.fileEntry.grid(columnspan=3, column=4, row=1, sticky=W)

		ttk.Label(fileFrame, text='Select Sheet:').grid(columnspan=3, row=2, sticky=E)
		self.sheetCBox = ttk.Combobox(fileFrame, textvariable=self.selectedSheet, width=30, state='readonly')
		self.sheetCBox['values'] = ''
		self.sheetCBox.grid(columnspan=3, column=4, row=2, sticky=W)

		self.loadBtn = ttk.Button(fileFrame, text='Load File', style='loadBtn.TButton')
		self.loadBtn.state(['disabled'])
		self.loadBtn.grid(columnspan=3, row=3, sticky='W S')
		ttk.Label(fileFrame, textvariable=self.fileDisp).grid(columnspan=3, column=4, row=3, sticky='W S')

		for child in fileFrame.winfo_children(): child.grid_configure(padx=5, pady=10)


	def eventPass(self, event):
		self.askDir()


	def askDir(self):
		# Open ask directory dialog and set path
		self.filename = filedialog.askopenfilename(**self.fileOpt)
		if self.filename:
			FileSelection.filePath.set(self.filename)
			root.bind('<Return>', self.loadFile)
			self.loadBtn.bind('<Button-1>', self.loadFile)
			self.loadBtn.state(['!disabled'])
			self.loadBtn.focus_set()
			self.loadWorkbook()
		style.configure('fileBtn.TButton', relief=RAISED)


	def loadWorkbook(self):
	# Load workbook and setup sheet selection
		try:
			FileSelection.wb = openpyxl.load_workbook(FileSelection.filePath.get())
			self.sheetCBox['values'] = FileSelection.wb.get_sheet_names()
			self.sheetCBox.current(self.sheetCBox['values'].index(FileSelection.wb.active.title))
		except FileNotFoundError:
			self.fileDisp.set('Could not load file at ' + self.filename)


	def loadFile(self, event):
		# Load selected sheet from workbook
		try:
			FileSelection.sheet = FileSelection.wb.get_sheet_by_name(self.selectedSheet.get())
			self.fileDisp.set('Successfully loaded ' + str(self.selectedSheet.get()) + '.')
			ParamSelection.paramFrame.grid()
			Search.searchFrame.grid()
		except KeyError:
			self.fileDisp.set('Error loading ' + str(self.selectedSheet.get()) + '!')
		

############################################################################################################################
class SearchSelection:
# Search Selection Frame (Lower left)
	def __init__(self):
		# Frame setup
		sModeFrame = ttk.LabelFrame(root, text='Search Type: ', padding='3 3 12 12')
		sModeFrame.grid(columnspan=6, row=2, pady=10, sticky='N S W E')
		
		# Required variables
		self.searchMode = StringVar()
		self.searchMode.set('keyword')
		self.matchMode = StringVar()
		self.matchMode.set('offset')
		self.radioSet()

		# Interface elements
		ttk.Label(sModeFrame, text='Select Search Mode:').grid(columnspan=6, row=1, sticky=W)
		self.keywordRBtn = ttk.Radiobutton(sModeFrame, text='Keyword Variations', variable=self.searchMode, value='keyword', command=self.radioSet)
		self.keywordRBtn.grid(columnspan=3, row=2, sticky=W)
		self.exactRBtn = ttk.Radiobutton(sModeFrame, text='Exact Match', variable=self.searchMode, value='exact', command=self.radioSet)
		self.exactRBtn.grid(columnspan=3, column=4, row=2, sticky=W)

		ttk.Label(sModeFrame, text='').grid(column=0, row=3, sticky=(W, E)) # Divider
		ttk.Label(sModeFrame, text='Select Match Mode:').grid(columnspan=6, row=4, sticky=W)
		self.columnRBtn = ttk.Radiobutton(sModeFrame, text='Matched Pattern', variable=self.matchMode, value='offset', command=self.radioSet)
		self.columnRBtn.grid(columnspan=3, row=5, sticky=W)
		self.offsetRBtn = ttk.Radiobutton(sModeFrame, text='From Another Column', variable=self.matchMode, value='column', command=self.radioSet)
		self.offsetRBtn.grid(columnspan=3, column=4, row=5, sticky=W)

		for child in sModeFrame.winfo_children(): child.grid_configure(padx=5, pady=0)
		

	def radioSet(self):
		search = self.searchMode.get()
		match = self.matchMode.get()
	

############################################################################################################################
class ParamSelection:
# Parameter selection frame (Upper right)
	def __init__(self):
		### Frame setup ###
		ParamSelection.paramFrame = ttk.LabelFrame(root, text='Search Options: ', padding='3 3 12 12')
		ParamSelection.paramFrame.grid(columnspan=6, pady=10, row=3, sticky='N W S E')

		# Required variables
		ParamSelection.searchCol = StringVar() # Holds name of column to be searched
		ParamSelection.pasteCol = StringVar() # Holds name of column to paste data into
		ParamSelection.offsetMode = StringVar() # Currently select offset mode (Radio button)
		ParamSelection.offsetPattern = StringVar() # Holds text from pattern entry field
		ParamSelection.vcmd = ParamSelection.paramFrame.register(self.updateHandler) # Validation binding
		self.offsetPtrnLbl = StringVar() # Holds text of label above pattern entry

		# Set defaults
		ParamSelection.searchCol.set('Column: A to XFD')
		ParamSelection.pasteCol.set('Column: A to XFD')
		ParamSelection.offsetMode.set('pattern')

		# Interface elements
		ttk.Label(ParamSelection.paramFrame, text='Which column would you like to search?').grid(columnspan=5, row=0, sticky=E)
		ParamSelection.sColEntry = ttk.Entry(ParamSelection.paramFrame, width=17, textvariable=ParamSelection.searchCol, foreground='grey', validate='all', validatecommand=(ParamSelection.vcmd, '%V', '%W', '%P'))
		ParamSelection.sColEntry.grid(columnspan=2, column=5, row=0, sticky=W)
		
		ttk.Label(ParamSelection.paramFrame, text='Which column would you like to copy the selected data to?').grid(columnspan=5, row=1, sticky=E)
		ParamSelection.pColEntry = ttk.Entry(ParamSelection.paramFrame, width=17, textvariable=ParamSelection.pasteCol, foreground='grey', validate='all', validatecommand=(ParamSelection.vcmd, '%V', '%W', '%P'))
		ParamSelection.pColEntry.grid(columnspan=2, column=5, row=1, sticky=W)

		ttk.Label(ParamSelection.paramFrame, text='').grid(columnspan=7, row=2, sticky=(W, E)) # Divider
		ttk.Label(ParamSelection.paramFrame, text='Select Offset Type:').grid(columnspan=3, row=3, sticky=W)
		ttk.Label(ParamSelection.paramFrame, text='Pattern Configuration:').grid(columnspan=5, column=2, row=3, sticky=W)

		self.patternRBtn = ttk.Radiobutton(ParamSelection.paramFrame, text='Pattern', variable=ParamSelection.offsetMode, value='pattern', command=self.radioSet)
		self.patternRBtn.grid(column=0, row=4, sticky=W)
		self.charRBtn = ttk.Radiobutton(ParamSelection.paramFrame, text='Character Count', variable=ParamSelection.offsetMode, value='char', command=self.radioSet)
		self.charRBtn.grid(column=1, row=4, sticky=W)
		self.configBtn = ttk.Button(ParamSelection.paramFrame, text='Set Up Search Pattern', style='configBtn.TButton', command=self.clickConfigure)
		self.configBtn.grid(columnspan=5, column=2, row=4, sticky=W)
		self.ptrnEntry = ttk.Entry(ParamSelection.paramFrame, width=30, textvariable=ParamSelection.offsetPattern, validate='all', validatecommand=(ParamSelection.vcmd, '%V', '%W', '%P'))
		self.ptrnEntry.grid(columnspan=5, column=2, row=4, sticky=W)
		
		self.nameDict = {str(ParamSelection.sColEntry):{'textvar':ParamSelection.searchCol, 'placeholder':'Column: A to XFD', 'entryName':ParamSelection.sColEntry, 'type':'column'},
						 		   str(ParamSelection.pColEntry):{'textvar':ParamSelection.pasteCol, 'placeholder':'Column: A to XFD', 'entryName':ParamSelection.pColEntry, 'type':'column'},
						 		   str(self.ptrnEntry):{'textvar':ParamSelection.offsetPattern, 'placeholder':'Must be a number (Ex 10)', 'entryName':self.ptrnEntry, 'type':'pattern'},
								   'radioTriggerMapping':{'textvar':ParamSelection.offsetPattern, 'placeholder':'Must be a number (Ex 10)', 'entryName':self.ptrnEntry, 'type':'pattern'}}

		for child in ParamSelection.paramFrame.winfo_children(): child.grid_configure(padx=5, pady=5)
		self.configBtn.grid_configure(pady=2)
		self.ptrnEntry.grid_remove()
		ParamSelection.paramFrame.grid_remove()


	def radioSet(self):
		try:
			if ParamSelection.offsetMode.get() == 'pattern':
				self.ptrnEntry.grid_remove()
				self.configBtn.grid()
			else:
				self.configBtn.grid_remove()
				self.ptrnEntry.grid()
				self.updateHandler('radioChange', 'radioTriggerMapping', ParamSelection.offsetPattern.get())
		except KeyError:
			return


	def updateHandler(self, reason, varName, entryValue): 
	# Called on entry state change, decides where to pass task. Return True to allow edit, False to disallow
		if reason == 'radioChange': # Radio button was clicked, check new position
			if (not self.validateEntry(varName, entryValue) or ParamSelection.offsetPattern.get() == ''):
				self.setPlaceholder(varName, True)

		elif reason == 'focusin':
			self.remPlaceholder(varName)

		elif reason == 'focusout':
			self.setPlaceholder(varName, False)

		elif reason == 'key':
			if not self.validateEntry(varName, entryValue):
				return False

			elif (self.nameDict[varName]['type'] == 'pattern') and (ParamSelection.offsetMode.get() == 'char'):
				if entryValue != '':
					self.toggleEnable('en')
				else:
					self.toggleEnable('dis')
		return True		


	def validateEntry(self, varName, curEntryVal):
	# Validates the entry based on entry type. Returns True if pass, False if fail
		if self.nameDict[varName]['type'] == 'column':
			if not (curEntryVal.isalpha() or curEntryVal == ''):
				return False
			elif curEntryVal != '':
				try:
					column_index_from_string(str(curEntryVal).upper())
				except ValueError:
					return False

		elif ParamSelection.offsetMode.get() == 'char':
			if not (curEntryVal.isdigit() or curEntryVal == ''):
				return False

		return True


	def setPlaceholder(self, varName, forceSet):
	# Sets the placeholder text of the entry. Can be forced to override current text
		textvar = self.nameDict[varName]['textvar']

		if forceSet: # If force flag is set, override current value
			textvar.set(self.nameDict[varName]['placeholder'])
			self.nameDict[varName][str('entryName')].configure(foreground='grey')

		elif textvar.get() == '': # Check if Entry is empty before setting value
			if (self.nameDict[varName]['type'] == 'pattern' and ParamSelection.offsetMode.get() == 'char'): # If pattern entry, make sure char is selected
				textvar.set(self.nameDict[varName]['placeholder'])
				self.nameDict[varName][str('entryName')].configure(foreground='grey')

			elif self.nameDict[varName]['type'] == 'column': # Column entry
				textvar.set(self.nameDict[varName]['placeholder'])
				self.nameDict[varName][str('entryName')].configure(foreground='grey')


	def remPlaceholder(self, varName):
	# Removes the placeholder text of the entry
		textvar = self.nameDict[varName]['textvar']

		if textvar.get() == self.nameDict[varName]['placeholder']:
			textvar.set('')
			self.nameDict[varName][str('entryName')].configure(foreground='black')


	def toggleEnable(self, state):
		if state == 'en':
			Search.termEntry.configure(state='enabled')
			Search.searchOptionsBtn.configure(state='enabled')
			Search.permutBtn.configure(state='enabled')
		else:
			Search.termEntry.configure(state='disabled')
			Search.searchOptionsBtn.configure(state='disabled')
			Search.permutBtn.configure(state='disabled')


	def clickConfigure(self):
		PatternDialog()


############################################################################################################################
class PatternDialog():
	def __init__(self):
		try:
			# Restore existing window if it exists
			PatternDialog.toplevel.deiconify()

		except (AttributeError, TclError) as e:
			# Window does not exist. Create and center the toplevel window
			PatternDialog.toplevel = Toplevel()
			PatternDialog.toplevel.title('Pattern Search Configuration')
			# Create icon from base64 code
			Base64IconGen(PatternDialog.toplevel)

			PatternDialog.rowID = 0 # Counter to assign unique ID to each row
			PatternDialog.startAfterMatchValue = True

			# Create dictionary for dialog drop downs
			PatternDialog.valuesDict = {'typeCB':['Any Character', 'Non letter/space', 'Non digit/space', 'Letter', 'Digit', 'Specify Pattern'],
							   			'repeatCB':['Repeated', 'Repeat Until'],
							   			'terminateCB':['Space Character', 'Alphanumeric', 'Letter', 'Digit'],
							   			'joinCB':['Then', 'Or']}

			# Create subframes to store various widgets
			PatternDialog.titleFrame = ttk.Frame(PatternDialog.toplevel)
			PatternDialog.ruleFrame = ttk.Frame(PatternDialog.toplevel)
			PatternDialog.buttonFrame = ttk.Frame(PatternDialog.toplevel)

			# Pack subframes to top, mid, and bottom
			PatternDialog.titleFrame.pack(side=TOP, fill=X, expand=True)
			PatternDialog.ruleFrame.pack(fill=X, expand=True)
			PatternDialog.buttonFrame.pack(side=BOTTOM, fill=X, expand=True)

			# Add instruction label 
			ttk.Label(PatternDialog.titleFrame, text='Match the following rules:').pack(side=TOP, anchor=W)

			# Draw a rule selection row
			RuleDialog()

			### Draw bottom buttons ###
			cancelBtn = ttk.Button(PatternDialog.buttonFrame, text='Cancel', command=self.cancelDialog, style='cancelBtn.TButton')
			cancelBtn.pack(side=LEFT)

			startAfterMatchCB = ttk.Checkbutton(PatternDialog.buttonFrame, text='Start Matching After Search Term', variable=self.startAfterMatchValue, onvalue=True, offvalue=False)
			startAfterMatchCB.pack(side=LEFT)
			startAfterMatchCB.invoke()

			doneBtn = ttk.Button(PatternDialog.buttonFrame, text='Done', command=self.doneDialog, style='doneBtn.TButton')
			doneBtn.pack(side=RIGHT)

		# Get screen dimensions
		self.rX = root.winfo_rootx()
		self.rY = root.winfo_rooty()
		self.rHeight = root.winfo_height()
		self.rWidth = root.winfo_width()

		# Add padding to all items in the frame
		for child in PatternDialog.toplevel.winfo_children(): child.pack_configure(padx=5, pady=5)
		for child in PatternDialog.buttonFrame.winfo_children(): child.pack_configure(padx=5, pady=5)		

		# Move windows to center of parent frame
		root.update_idletasks()
		size = list(int(item) for item in PatternDialog.toplevel.geometry().split('+')[0].split('x'))
		geometry = "+%d+%d" % (self.rX + ((self.rWidth / 2) - (size[0] / 2)), self.rY + ((self.rHeight / 2) - (size[1] / 2)))
		PatternDialog.toplevel.geometry(geometry)

		# Everything appears to have gone well, lock the parent frame while top level is active
		PatternDialog.toplevel.grab_set()


	def cancelDialog(self):
		ParamSelection.offsetPattern.set('')
		PatternDialog.toplevel.destroy()
		PatternDialog.toplevel.grab_release()
		PatternDialog.instanceDict = {}
		RegexGeneration.rulesDict = {}
		ParamSelection.toggleEnable(ParamSelection, 'dis')


	def doneDialog(self):
		PatternDialog.toplevel.withdraw()
		PatternDialog.toplevel.grab_release()
		ParamSelection.toggleEnable(ParamSelection, 'en')


############################################################################################################################
class RuleDialog:
# Class to handle creation and destruction of rules rows
	def __init__(self):
		# Name this instance and add it to rule dictionary
		self.name = 'row ' + str(PatternDialog.rowID)
		PatternDialog.rowID += 1
		RegexGeneration.rulesDict[self.name] = [''] * 7
		RegexGeneration.rulesDict[self.name][0] = (self)

		# Required variables
		self.typeValue = StringVar()
		self.repeatValue = StringVar()
		self.terminateValue = StringVar()
		self.joinValue = StringVar()
		self.charEntryValue = StringVar()
		self.charEntryValue.trace('w', self.updateDict)
		self.repeatEntryValue = StringVar()
		self.repeatEntryValue.trace('w', self.updateDict)

		# Create inner frame for layout and pack it
		self.innerFrame = ttk.Frame(PatternDialog.ruleFrame, style='innerFrame.TFrame')
		self.innerFrame.pack(anchor=W)

		# 1st rule section button, type of character (or exact char) to match
		self.typeCB = ttk.Combobox(self.innerFrame, textvariable=self.typeValue, width=16, state='readonly')
		self.typeCB['values'] = PatternDialog.valuesDict['typeCB']
		self.typeCB.current(0)
		self.typeCB.bind('<<ComboboxSelected>>', self.valueChanged)
		self.typeCB.pack(side=LEFT, anchor=W, padx=2, pady=5)

		# Placeholder frame
		self.optionFrameOne = ttk.Frame(self.innerFrame, style='optionFrame.TFrame')
		self.optionFrameOne.pack(side=LEFT, anchor=W)

		# Optional space to specify character
		self.charEntry = ttk.Entry(self.optionFrameOne, width=10, textvariable=self.charEntryValue)
		self.charEntry.pack(side=LEFT, anchor=W, padx=5, pady=5)

		# 2nd rule section button, how it should be allowed to repeat
		self.repeatCB = ttk.Combobox(self.innerFrame, textvariable=self.repeatValue, width=11, state='readonly')
		self.repeatCB['values'] = PatternDialog.valuesDict['repeatCB']
		self.repeatCB.current(1)
		self.repeatCB.bind('<<ComboboxSelected>>', self.valueChanged)
		self.repeatCB.pack(side=LEFT, anchor=W, padx=5, pady=5)

		# Placeholder frame
		self.optionFrameTwo = ttk.Frame(self.innerFrame, style='optionFrame.TFrame')
		self.optionFrameTwo.pack(side=LEFT, anchor=W)

		# Optional entry to specify occurrences
		self.repeatEntry = ttk.Entry(self.optionFrameTwo, width = 3, textvariable=self.repeatEntryValue)
		self.repeatEntry.pack(side=LEFT, anchor=W, padx=2, pady=5)
		self.repeatLbl = ttk.Label(self.optionFrameTwo, text='times')
		self.repeatLbl.pack(side=LEFT, anchor=W, padx=2, pady=5)

		# 3rd rule section button, repeat termination (if repeat until is selected)
		self.terminateCB = ttk.Combobox(self.optionFrameTwo, textvariable=self.terminateValue, width=19, state='readonly')
		self.terminateCB['values'] = PatternDialog.valuesDict['terminateCB']
		self.terminateCB.current(0)
		self.terminateCB.bind('<<ComboboxSelected>>', self.valueChanged)
		self.terminateCB.pack(side=LEFT, anchor=W, padx=5, pady=5)

		# Placeholder frame
		self.joinFrame = ttk.Frame(self.innerFrame, style='optionFrame.TFrame')
		self.joinFrame.pack(side=LEFT, anchor=W)

		# Optionally add a button to toggle AND / OR (if there is more than 1 row)
		self.joinCB = ttk.Combobox(self.joinFrame, textvariable=self.joinValue, width=5, state='readonly')
		self.joinCB['values'] = PatternDialog.valuesDict['joinCB']
		self.joinCB.current(0)
		self.joinCB.bind('<<ComboboxSelected>>', self.valueChanged)
		self.joinCB.pack(side=LEFT, anchor=W, padx=5, pady=5)
		self.joinCB.pack_forget()

		# - button to remove rule
		self.removeBtn = ttk.Button(self.innerFrame, text = '-', command=self.removeRule, style='removeBtn.TButton')
		self.removeBtn.config(width=3)
		self.removeBtn.pack(side=LEFT, anchor=E, padx=5, pady=5)

		# + button to add another rule
		self.addBtn = ttk.Button(self.innerFrame, text = '+', command=self.addRule, style='addBtn.TButton')
		self.addBtn.config(width=3)
		self.addBtn.pack(side=LEFT, anchor=E, padx=5, pady=5)

		# Add values to dictionary
		self.valueChanged(None)


	def valueChanged(self, event):
		self.updateDict(None, None, None)
		self.updateDisplay()


	def updateDisplay(self):
		for ID, value in RegexGeneration.rulesDict.items():
			if 'Specify Character' in value:
				value[0].optionFrameTwo.pack_propagate(True)
				value[0].charEntry.pack(side=LEFT, anchor=W, padx=5, pady=5)
			else:
				value[0].charEntry.pack_forget()
				value[0].optionFrameOne.configure(width=1, height=1)

			if 'Repeated' in value:
				value[0].optionFrameTwo.pack_propagate(True)
				value[0].repeatEntry.pack(side=LEFT, anchor=W, padx=2, pady=5)
				value[0].repeatLbl.pack(side=LEFT, anchor=W, padx=2, pady=5)
				value[0].terminateCB.pack_forget()
			else:
				value[0].repeatEntry.pack_forget()
				value[0].repeatLbl.pack_forget()
				value[0].optionFrameTwo.configure(width=1, height=1)
				value[0].terminateCB.pack(side=LEFT, anchor=W, padx=5, pady=5)


	def lineConcatUpdate(self):
		# Find the max and min row numbers
		rowList = []
		for rowNum in RegexGeneration.rulesDict.keys():
			# Extract number from name
			rowList.append(int(rowNum.split()[1]))
			# Find max row
			maxRow = max(rowList)
			# Find min row
			minRow = min(rowList)

		if len(RegexGeneration.rulesDict.keys()) > 1:
			for rowID, value in RegexGeneration.rulesDict.items():
				# Add line concatenation operators
				value[0].joinFrame.pack_propagate(True)
				value[0].joinCB.pack(side=LEFT, anchor=W, padx=5, pady=5)

				# If this is the 2nd item added, also update the first row
				if len(RegexGeneration.rulesDict.keys()) < 3:
					value[0].joinFrame.pack_propagate(True)
					value[0].joinCB.pack(side=LEFT, anchor=W, padx=5, pady=5)

				# If this is the last row, omit the concatenation operator selection box
				if int(rowID.split()[1]) == maxRow:
					RegexGeneration.rulesDict['row ' + str(maxRow)][0].joinFrame.configure(width=1, height=1)
					RegexGeneration.rulesDict['row ' + str(maxRow)][0].joinCB.pack_forget()
		else:
			RegexGeneration.rulesDict['row ' + str(maxRow)][0].joinFrame.configure(width=1, height=1)
			RegexGeneration.rulesDict['row ' + str(maxRow)][0].joinCB.pack_forget()


	def updateDict(self, name, index, mode):
		print('Self: ' + str(self) + ' name: ' + str(name) + ' index: ' + str(index) + ' mode: ' + str(mode))
		RegexGeneration.rulesDict[self.name][1] = str(self.typeValue.get())
		RegexGeneration.rulesDict[self.name][2] = str(self.repeatValue.get())
		RegexGeneration.rulesDict[self.name][3] = str(self.terminateValue.get())
		RegexGeneration.rulesDict[self.name][4] = str(self.charEntryValue.get())
		RegexGeneration.rulesDict[self.name][5] = str(self.repeatEntryValue.get())
		print('Char: ' + str(self.repeatEntryValue.get()))
		RegexGeneration.rulesDict[self.name][6] = str(self.joinValue.get())
		print(RegexGeneration.rulesDict[self.name])


	def removeRule(self):
		if len(RegexGeneration.rulesDict.keys()) > 1:
			RegexGeneration.rulesDict.pop(self.name)
			self.innerFrame.destroy()
			self.lineConcatUpdate()


	def addRule(self):
		RuleDialog()
		self.lineConcatUpdate()


############################################################################################################################
class Search:
# Search Frame (Lower right)
	def __init__(self):
		### Outer Frame setup ###
		Search.searchFrame = ttk.LabelFrame(root, text='Search: ', padding='3 3')
		Search.searchFrame.grid(columnspan=6, column=7, rowspan=4, row=0, pady=10, sticky='N W S E')

		# Required variables
		self.searchTerm = StringVar()
		self.selectStateText = StringVar()
		self.selectStateText.set('Select All')
		self.cbVals={}
		self.resultCB={}
		self.lbl={}
		
		# Interface elements
		ttk.Label(Search.searchFrame, text='Enter search term:').grid(columnspan=2, row=2, sticky='W')
		Search.termEntry = ttk.Entry(Search.searchFrame, width=53, textvariable=self.searchTerm)
		Search.termEntry.bind('<FocusIn>', self.createHandler)
		Search.termEntry.grid(columnspan=5, column=2, row=2, sticky=W)
		Search.termEntry.configure(state='disabled')
		
		Search.searchOptionsBtn = ttk.Button(Search.searchFrame, text='Search Options', command=self.optionsDialog, style='searchOptionsBtn.TButton')
		Search.searchOptionsBtn.grid(column=0, row=3, sticky='W')
		self.instructionLbl = ttk.Label(Search.searchFrame, text="A period will allow any non alphanumeric character to be substituted in it's place.", wraplength=350)
		self.instructionLbl.grid(columnspan=6, column=1, row=3, sticky='W')
		Search.searchOptionsBtn.configure(state='disabled')

		specialLbl = ttk.Label(Search.searchFrame, text='Available Permutations:').grid(columnspan=3, row=4, sticky='W S')
		Search.permutBtn = ttk.Button(Search.searchFrame, text='Search Permutations', command=self.searchPerms, style='permutBtn.TButton')
		Search.permutBtn.grid(columnspan=2, column=5, row=4, sticky=E)
		Search.permutBtn.configure(state='disabled')

		Search.selectBtn = ttk.Button(Search.searchFrame, textvariable=self.selectStateText, command=self.switchState, style='selectBtn.TButton')
		Search.selectBtn.grid(row=9, sticky=W)
		Search.startSearchBtn = ttk.Button(Search.searchFrame, text='Start Search', command=self.startSearch, style='startSearchBtn.TButton')
		Search.startSearchBtn.grid(columnspan=5, column=1, row=9)
		Search.exportBtn = ttk.Button(Search.searchFrame, text='Export Sheet', command=self.exportSheet, style='exportBtn.TButton')
		Search.exportBtn.grid(columnspan=2, column=6, row=9, sticky=E)
		Search.startSearchBtn.configure(state='disabled')
		Search.selectBtn.configure(state='disabled')
		Search.exportBtn.configure(state='disabled')

		for child in Search.searchFrame.winfo_children(): child.grid_configure(padx=5, pady=10)
		Search.searchOptionsBtn.grid_configure(pady=0)
		self.instructionLbl.grid_configure(padx=7, pady=0)

		### Inner Frame Setup ###
		self.resultFrameTop = ttk.Frame(Search.searchFrame, borderwidth=0, relief='groove', padding='3 3 170 120')
		self.resultFrameTop.grid(columnspan=7, row=5, padx=5, sticky='N W S E')
		self.resultFrameBot = ttk.Frame(Search.searchFrame, borderwidth=0)
		self.resultFrameBot.grid(columnspan=7, rowspan=2, row=6, padx=5, sticky='N W S E')

		# Required Variables
		self.results = StringVar()
		self.results.set('Waiting for search term...')
		Search.drawCalled = False

		# Interface elements
		ttk.Label(self.resultFrameTop, textvariable=self.results).pack(side=LEFT)
		Search.searchFrame.grid_remove()


	def createHandler(self, event):
		ExcelHandler(FileSelection, ParamSelection)

	
	def drawResults(self):
		# Setup result headings
		self.results.set('')
		self.resultFrameTop.configure(relief='flat', padding='0 0 0 0')
		self.resultFrameTop.grid_configure(padx=1)
		self.resultHLbl = ttk.Label(self.resultFrameTop, text='Result:' + ' ' * 25)
		self.resultHLbl.pack(side=LEFT)
		self.posHLbl = ttk.Label(self.resultFrameTop, text='Row/Col:' + ' ' * 10)
		self.posHLbl.pack(side=LEFT)
		self.contextHLbl = ttk.Label(self.resultFrameTop, text='Context:')
		self.contextHLbl.pack(side=LEFT)

		# Set up frame
		self.scrollbar = Scrollbar(self.resultFrameBot, orient=VERTICAL)
		self.resultbox = Listbox(self.resultFrameBot, selectmode=MULTIPLE, yscrollcommand=self.scrollbar.set)
		self.scrollbar.config(command=self.resultbox.yview)
		self.scrollbar.pack(side=RIGHT, fill=Y, padx=0, pady=0)
		self.resultbox.pack(fill=BOTH, expand=1, padx=0, pady=0)

		# Remove all items from previous list
		self.resultbox.delete(1, END)
		# Set called var
		Search.drawCalled = True


	def optionsDialog(self):
		print('Options!')


	def searchPerms(self):
	# Search the selected spreadsheet for the specified permutations and generate any requested ones
		# Check if the user actually entered something
		if self.searchTerm.get() != '':
			# Enable buttons
			Search.startSearchBtn.configure(state='enabled')
			Search.selectBtn.configure(state='enabled')

			# Draw result box if it's not already there
			if not Search.drawCalled:
				self.drawResults()
			
			# Remove all items from previous list
			self.resultbox.delete(0, END)

			# Generate permutations based on search terms
			RegexGeneration.genPerms(RegexGeneration, str(self.searchTerm.get()))
			ExcelHandler.findPerms(ExcelHandler)
			
			# Print results
			for item in permsFound:
				if not (item in self.resultbox.get(0, "end")):
					self.resultbox.insert(END, item)
		else:
			tkinter.messagebox.showerror('Error', 'Please enter a term to search for.')


	def switchState(self):
		if self.selectStateText.get() == 'Deselect All':
			self.selectStateText.set('Select All')
			self.resultbox.selection_clear(0, END)
		else:
			self.selectStateText.set('Deselect All')
			self.resultbox.selection_set(0, END)


	def startSearch(self):
		self.translateSelection()
		ExcelHandler.searchSheet(ExcelHandler)
		Search.exportBtn.configure(state='enabled')
		tkinter.messagebox.showinfo('Progress', 'Finished Search.')


	def translateSelection(self):
		# Correlates the currently selected listbox items to their string values using the permsFound list
		global permsToSearch
		permsToSearch = self.resultbox.curselection()
		permsToSearch = [permsFound[int(item)] for item in permsToSearch]


	def exportSheet(self):
		exportPath = os.path.dirname(ExcelHandler.filePath) + os.sep + 'extracted_' + os.path.basename(ExcelHandler.filePath)
		ExcelHandler.wb.save(exportPath)
		text = 'Exported to ' + exportPath + '\n\nStart another extraction?'
		ans = tkinter.messagebox.askquestion('Export', text)

		if ans == 'yes':
			ObjectManagement.reset()


############################################################################################################################
class ExcelHandler(FileSelection, ParamSelection):
# Class to handle traversing Excel sheets
	def __init__(self, files, params):
		ExcelHandler.filePath = files.filePath.get()
		ExcelHandler.wb = files.wb
		ExcelHandler.sheet = files.sheet
		try:
			ExcelHandler.searchCol = column_index_from_string(str(params.searchCol.get()).upper())
			ExcelHandler.pasteCol = column_index_from_string(str(params.pasteCol.get()).upper())
		except ValueError:
			ParamSelection.sColEntry.focus_set()
			tkinter.messagebox.showerror('Error', 'Please enter valid column names.')
		

	def findPerms(self):
		global permsFound
		for rowNum in range (1, ExcelHandler.sheet.max_row + 1):
			curCell = ExcelHandler.sheet.cell(row=rowNum, column=ExcelHandler.searchCol)
			for result in RegexGeneration.permutRegex.findall(str(curCell.value)):
				if result[0] not in (' '.join(permsFound)):
					permsFound.append(result[0])


	def getPermInfo(self):
		print('Placeholder for getting Row/Col and Context info')


	def searchSheet(self):
		for term in permsToSearch:
			RegexGeneration.parsePattern(RegexGeneration, term)
			for rowNum in range (1, ExcelHandler.sheet.max_row + 1):
				curCell = ExcelHandler.sheet.cell(row=rowNum, column=ExcelHandler.searchCol)
				startIndex = str(curCell.value).find(term)
				if startIndex != -1:
					print('Found match for ' + term + ' at ' + str(curCell)) #Eventually change to live feedback popup (scrolls with this text)
					self.matchItem(self, curCell, ExcelHandler.sheet.cell(row=rowNum, column=ExcelHandler.pasteCol))


	def matchItem(self, searchCell, pasteCell):
		try:
			result = RegexGeneration.searchPatternRegex.search(searchCell.value)
			if result != None:
				pasteValue = result.group()
				for char in RegexGeneration.charsToRemove:
					pasteValue = pasteValue.replace(char, '')
				pasteCell.value = pasteValue
		except TypeError:
			print('TypeError!')
			return


############################################################################################################################
class RegexGeneration:
# Class to handle all regular expression and pattern generation

	def __init__(self):
		# Set up dictionary for permutation matching
		RegexGeneration.rulesDict = {}
		RegexGeneration.charsToRemove = ': '


	def genPerms(self, originTerm):
	# Generate permutations to search for based on search term
		if '.' not in originTerm:
			searchStrings = originTerm.split()
			searchStrings = '([\\-_ /])*'.join(searchStrings)
			RegexGeneration.permutRegex = re.compile(r'(' + searchStrings + ')+', re.I)
			RegexGeneration.generatedPerm = searchStrings

		else:
			searchStrings = originTerm.split()
			searchStrings = '([\\-_ /])*'.join(searchStrings)
			searchStrings = originTerm.split('.')
			searchStrings = '([\\-_ /])*'.join(searchStrings)
			RegexGeneration.generatedPerm = searchStrings
			RegexGeneration.permutRegex = re.compile(r'(' + searchStrings + ')+', re.I)


	def parsePattern(self, currentTerm):
	# Parse pattern generated by pattern dialog box
		RegexGeneration.initialSearchPattern = '' # Clear (to rebuild) global pattern each time Done is clicked

		# If mode is char, skip pattern gen
		if ParamSelection.offsetMode.get() == 'char':
			RegexGeneration.initialSearchPattern = '(?<=' + currentTerm + ')([:-_=,. ]*\S{' + str(ParamSelection.offsetPattern.get()) + '})'
			print(RegexGeneration.initialSearchPattern)

		# Mode is pattern, generate pattern
		else:
			for row in sorted(RegexGeneration.rulesDict.keys()): # Got row by row through pattern dialog settings
				setting = RegexGeneration.rulesDict[row] # store setting values for this row into more readable variable

				pattern = '' # Clear row pattern each row iteration

				# Evaluate tybeCB field
				if setting[1] == 'Any Character':
					pattern += '.'
				elif setting[1] == 'Non letter/space':
					pattern += '[^a-zA-Z ]'
				elif setting[1] == 'Non digit/space':
					pattern += '[^\d ]'
				elif setting[1] == 'Letter':
					pattern += '[a-zA-z]'
				elif setting[1] == 'Digit':
					pattern += '\d'
				elif setting[1] == 'Specify Character':
					pattern += str(setting[4])

				terminator = '' # Clear pattern terminator each row iteration

				# Check if pattern will repeat x times or if it needs to look for a pattern
				if setting[2] == 'Repeat Until':
					# Looking for pattern, evaluate terminateCB field
					if setting[3] == 'Space Character':
						terminator = '\s'
					elif setting[3] == 'Aplhanumeric':
						terminator = '\w'
					elif setting[3] == 'Letter':
						terminator = '[a-zA-z]'
					elif setting[3] == 'Digit':
						terminator = '\d'

				# Evaluate repeatCB field
				if setting[2] == 'Repeated':
					pattern = '(' + pattern + '){' + str(setting[5]) + '}' # Repeat pattern x times
				elif setting[2] == 'Repeat Until':
					pattern = '(' + pattern + ')+?' + '(?=' + terminator + ')|(' + pattern + ')+?' + '(?=$)' # Look for terminator pattern and pattern at end of cell
				
				# Evaluate joinCB field
				if len(RegexGeneration.rulesDict.keys()) > 1: 
				# There is more than one row in the pattern gen dialog box, join lines

					if setting[6] == 'Then':
					# Rows should be one continuous pattern
						pattern = '(' + pattern + ')+' # Make a group and look for one or more instances of it

						# Check where to start looking for a match
						if PatternDialog.startAfterMatchValue:
							pattern = '(?<=' + currentTerm + '[:-_=,. ])([:-_=,. ])*' + pattern # Start matching after current term permutation is found. Ignore all [:-_=,. ] until until match is found

						RegexGeneration.initialSearchPattern += pattern # Append row pattern to total pattern

					if setting[6] == 'Or':
					# Rows should be individual patterns ORed together

						# Check where to start looking for a match
						if PatternDialog.startAfterMatchValue:
						# Start matching after current term permutation is found. Ignore all [:-_=,. ] until until match is found
							pattern = '(?<=' + currentTerm + '[:-_=,. ])([:-_=,. ])*' + '(' + pattern + ')' + '+|' # Make group and search for one or more instance of it. Add | to OR it with next line
						else:
							pattern = '(' + pattern + ')' + '+|' # Make group and search for one or more instance of it. Add | to OR it with next line (start matching from beginning)

						RegexGeneration.initialSearchPattern += pattern # Append row pattern to total pattern

				elif len(RegexGeneration.rulesDict.keys()) == 1: # There is one row in the pattern gen dialog box

					# Check where to start looking for a match
					if PatternDialog.startAfterMatchValue:
					# Start matching after current term permutation is found. Ignore all [:-_=,. ] until until match is found
						pattern = '(?<=' + currentTerm + '[:-_=,. ])([:-_=,. ])*' + '(' + pattern + ')+'
					else:
						pattern = '(' + pattern + ')+' # Make group and search for one or more instance of it

					RegexGeneration.initialSearchPattern = pattern # Set row pattern to final pattern

		# Actually generate the pattern
		RegexGeneration.searchPatternRegex = re.compile(RegexGeneration.initialSearchPattern, re.I)


############################################################################################################################
class Base64IconGen():
# Takes icon from base64 format and creates window icon
	def __init__(self, window):
		icondata = base64.b64decode(base64ico.extractorIcon)
		# The temp file is icon.ico
		tempFile= "icon.ico"
		iconfile= open(tempFile,"wb")
		# Extract the icon
		iconfile.write(icondata)
		iconfile.close()
		window.wm_iconbitmap(tempFile)
		# Delete the tempfile
		os.remove(tempFile)


############################################################################################################################
class ObjectManagement():
# Manages creation and deletion of object classes

	def createObjects():
		ObjectManagement.regex =  RegexGeneration()
		ObjectManagement.filePane = FileSelection()
		ObjectManagement.seachModePane = SearchSelection()
		ObjectManagement.paramPane = ParamSelection()
		ObjectManagement.searchPane = Search()


	def deleteObjects():
		del ObjectManagement.regex
		del ObjectManagement.filePane
		del ObjectManagement.seachModePane
		del ObjectManagement.paramPane
		del ObjectManagement.searchPane


	def reset():
		global root
		ObjectManagement.deleteObjects() # Delete all objects

		# Reset root window
		root.destroy()
		root = Tk() # Create blank window
		root.title('Excel Extractor') # Set the name
		Base64IconGen(root)

		ObjectManagement.createObjects() # Create new objects


#************************************ Program Start ************************************#
# Create objects
ObjectManagement.createObjects()

# Create icon
Base64IconGen(root)

# Run GUI
root.mainloop()