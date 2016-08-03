#! python3

#########################################################################
# ExcelCellSwap.py 														#	
# Written by Kanyon Edvall 												#
# 																		#
# This program takes an Excel spreadsheet from accounting in horizontal	#
# format and converts it to vertical format for import to IFS. 			#
#########################################################################
# Import necessary modules
import openpyxl, os, sys, getpass, base64, base64ico
import tkinter.messagebox
from tkinter import *
from tkinter import ttk, filedialog
from inspect import currentframe
from openpyxl.cell import get_column_letter, column_index_from_string


# Set up GUI
root = Tk() # Create blank window
root.title('Excel Extractor') # Set the name
style = ttk.Style() # Set the style


# Class declaration
############################################################################################################################
class FileSelection:
# File Selection Frame (Upper left)

	def __init__(self):
		### Frame setup ###
		fileFrame = ttk.LabelFrame(root, text='File Selection: ', padding='3 3 12 12') # Make a themed frame to hold objects
		fileFrame.grid(columnspan=6, row=0, pady=10, sticky='N W S E')

		# Required variables
		FileSelection.filePath = StringVar()
		self.selectedSheet = StringVar()
		self.fileDisp = StringVar()

		# File opening options
		self.fileOpt = options = {}
		options['defaultextension'] = '.xlsx'
		options['filetypes'] = [('Excel', '.xlsx'), ('Spreadsheet', '.xlsm'), ('Spreadsheet', '.xltx'), ('Spreadsheet', '.xltm')]
		options['initialdir'] = 'C:\\Users\\' + getpass.getuser() + '\\Documents'
		options['parent'] = fileFrame
		options['title'] = 'Select a File to Use'

		# Interface elements
		self.fileBtn = ttk.Button(fileFrame, text='Select File:', style='fileBtn.TButton', command=self.askDir)
		root.bind('<Return>', self.eventPass)
		self.fileBtn.focus_set()
		self.fileBtn.grid(columnspan=3, row=1, sticky=E)
		self.fileEntry = ttk.Entry(fileFrame, width=50, textvariable=self.filePath)
		self.fileEntry.grid(columnspan=3, column=4, row=1, sticky=W)

		ttk.Label(fileFrame, text='Select Sheet:').grid(columnspan=3, row=2, sticky=E)
		self.sheetCBox = ttk.Combobox(fileFrame, textvariable=self.selectedSheet, width=30, state='readonly')
		self.sheetCBox['values'] = ''
		self.sheetCBox.grid(columnspan=3, column=4, row=2, sticky=W)

		self.loadBtn = ttk.Button(fileFrame, text='Load File', style='loadBtn.TButton')
		self.loadBtn.state(['disabled'])
		self.loadBtn.grid(columnspan=3, row=3, sticky='W S')
		ttk.Label(fileFrame, textvariable=self.fileDisp).grid(columnspan=3, column=4, row=3, sticky='W S')

		for child in fileFrame.winfo_children(): child.grid_configure(padx=5, pady=10)


	def eventPass(self, event):
		self.askDir()


	def askDir(self):
		# Open ask directory dialog and set path
		self.filename = filedialog.askopenfilename(**self.fileOpt)
		if self.filename:
			FileSelection.filePath.set(self.filename)
			root.bind('<Return>', self.loadFile)
			self.loadBtn.bind('<Button-1>', self.loadFile)
			self.loadBtn.state(['!disabled'])
			self.loadBtn.focus_set()
			self.loadWorkbook()
		style.configure('fileBtn.TButton', relief=RAISED)


	def loadWorkbook(self):
	# Load workbook and setup sheet selection
		try:
			FileSelection.wb = openpyxl.load_workbook(FileSelection.filePath.get())
			self.sheetCBox['values'] = FileSelection.wb.get_sheet_names()
			self.sheetCBox.current(self.sheetCBox['values'].index(FileSelection.wb.active.title))
		except FileNotFoundError:
			self.fileDisp.set('Could not load file at ' + self.filename)


	def loadFile(self, event):
		# Load selected sheet from workbook
		try:
			FileSelection.sheet = FileSelection.wb.get_sheet_by_name(self.selectedSheet.get())
			self.fileDisp.set('Successfully loaded ' + str(self.selectedSheet.get()) + '.')
			SwapValues.mainFrame.grid()
		except KeyError:
			self.fileDisp.set('Error loading ' + str(self.selectedSheet.get()) + '!')
		

class SwapValues():
# Handles the actual swapping of values

	def __init__(self):
	# Generate necessary settings
		# Matching rules - This dictionary defines what the various columns map to
		self.mappingRulesDict = {'A':'yearVal',
								 'B':'periodVal',
								 'C':'A',
								 'D':'B',
								 'E':'C',
								 'F':'D',
								 'G':None,
								 'H':None,
								 'I':'E',
								 'J':None,
								 'K':None,
								 'L':None,
								 'M':'budgetMap',
								 'N':None,
								 'O':None}

		# Budget Period to Column matching rules - This dictionary defines what period corresponds with what column
		self.periodRulesDict = {'1':column_index_from_string('G'),
							    '2':column_index_from_string('H'),
							    '3':column_index_from_string('I'),
							    '4':column_index_from_string('J'),
							    '5':column_index_from_string('K'),
							    '6':column_index_from_string('L'),
							    '7':column_index_from_string('M'),
							    '8':column_index_from_string('N'),
							    '9':column_index_from_string('O'),
							    '10':column_index_from_string('P'),
							    '11':column_index_from_string('Q'),
							    '12':column_index_from_string('R')}

		# Headers - This dictionary defines what headers to user for the new Excel sheet
		self.headersDict = {'A':None,
						    'B':'BUDGET_PERIOD',
						    'C':'ACCOUNT',
						    'D':'CODE_B',
						    'E':'CODE_C',
						    'F':'CODE_D',
						    'G':'CODE_E',
						    'H':'CODE_F',
						    'I':'CODE_G',
						    'J':'CODE_H',
						    'K':'CODE_I',
						    'L':'CODE_J',
						    'M':'AMOUNT',
						    'N':'QUANTITY',
						    'O':'TEXT'}

		# Other settings
		self.startRow = 2
		self.printRow = self.startRow

		self.setupFrame() # Configure frame


	def setupFrame(self):
		### Frame setup ###
		SwapValues.mainFrame = ttk.LabelFrame(root, text='Import Options: ', padding='3 3 12 12')
		SwapValues.mainFrame.grid(columnspan=6, pady=10, row=3, sticky='N W S E')

		### Required variables ###
		SwapValues.yearSV = StringVar() # Variable to hold entered column

		### Interface elements ###
		# Subframes to hold various elements
		SwapValues.entryFrame = ttk.Frame(SwapValues.mainFrame)
		SwapValues.entryFrame.pack(fill=X, expand=True)
		SwapValues.buttonFrame = ttk.Frame(SwapValues.mainFrame)
		SwapValues.buttonFrame.pack(side=BOTTOM, fill=X, expand=True)

		SwapValues.vcmd = SwapValues.mainFrame.register(self.validateYear) # Register validate command on new frame

		# Year entry field
		ttk.Label(SwapValues.entryFrame, text='Enter year to print to the Excel sheet:').grid(columnspan=4, row=0, sticky=W)
		SwapValues.yearEntry = ttk.Entry(SwapValues.entryFrame, width=5, textvariable=SwapValues.yearSV, validate='key', validatecommand=(SwapValues.vcmd, '%P'))
		SwapValues.yearEntry.grid(columnspan=2, column=4, row=0, sticky=W)

		# Create a settings button
		SwapValues.addButton = ttk.Button(SwapValues.buttonFrame, text='Swap Settings', command=self.showSettings)
		SwapValues.addButton.pack(side=LEFT, anchor=W, pady=5, padx=5)

		# Create Start Swap button
		SwapValues.importButton = ttk.Button(SwapValues.buttonFrame, text='Start Swap', command=self.swapValues)
		SwapValues.importButton.pack(side=RIGHT, anchor=E, pady=5, padx=5)
		SwapValues.importButton.configure(state='disabled')

		# Add padding for frames
		for child in SwapValues.entryFrame.winfo_children(): 
			child.grid_configure(padx=5, pady=5)

		for child in SwapValues.buttonFrame.winfo_children(): 
			child.pack_configure(padx=5, pady=5)

		SwapValues.mainFrame.grid_remove() # Temporarily hid this frame


	def validateYear(self, yearValue):
		if not (yearValue.isdigit() or yearValue == ''):
			SwapValues.importButton.configure(state='disabled')
			return False

		if yearValue != '':
			SwapValues.importButton.configure(state='enabled')

		return True


	def getVal(self, valueToMap, rowNum):
		if valueToMap == None:
			return None
		elif valueToMap == 'yearVal':
			return SwapValues.yearSV.get()
		elif valueToMap == 'periodVal':
			return self.budgetPeriod
		elif valueToMap == 'budgetMap':
			return FileSelection.sheet.cell(row=rowNum, column=self.periodRulesDict[str(self.budgetPeriod)]).value
		else:
			return FileSelection.sheet.cell(row=rowNum, column=column_index_from_string(valueToMap)).value


	def swapValues(self):
	# Swap columns based on matching rules
		# Check date
		if len(SwapValues.yearSV.get()) != 4:
			tkinter.messagebox.showerror('Date Error', 'Please enter the full date (Ex 2016 not 16)')

		# Perform swap if date format is correct
		else:
			# Prep workbook and sheet
			SwapValues.newWb = openpyxl.Workbook()
			SwapValues.newSheet = SwapValues.newWb.get_sheet_by_name('Sheet')
			SwapValues.newSheet.title = 'Swapped Sheet'

			# Format sheet with correct headers
			for key, value in sorted(self.headersDict.items()):
				SwapValues.newSheet.cell(row=self.startRow, column=column_index_from_string(key)).value = value

			# Pull data from matching defined place
			for rowNum in range(1, FileSelection.sheet.max_row):
				for self.budgetPeriod in range(1, 13):
					self.printRow += 1
					for key, value in sorted(self.mappingRulesDict.items()):
						SwapValues.newSheet.cell(row=self.printRow, column=column_index_from_string(key)).value = self.getVal(value, rowNum)

			# Export new workbook
			self.exportSheet()

	
	def exportSheet(self):
		exportPath = os.path.dirname(FileSelection.filePath.get()) + os.sep + 'swapped_' + os.path.basename(FileSelection.filePath.get())
		SwapValues.newWb.save(exportPath)
		text = 'Exported to ' + exportPath
		tkinter.messagebox.showinfo('Export', text)


	def showSettings(self):
		message = '\n'.join(['\t=\t'.join([key, str(val)]) for key, val in sorted(self.headersDict.items())])
		tkinter.messagebox.showinfo('Current header settings', message)

		message = '\n'.join(['\t->\t'.join([key, str(val)]) for key, val in sorted(self.mappingRulesDict.items())])
		tkinter.messagebox.showinfo('Current column mapping table', message)


############################################################################################################################
class Base64IconGen():
# Takes icon from base64 format and creates window icon
	def __init__(self, window):
		icondata = base64.b64decode(base64ico.swapIcon)
		# The temp file is icon.ico
		tempFile= "icon.ico"
		iconfile= open(tempFile,"wb")
		# Extract the icon
		iconfile.write(icondata)
		iconfile.close()
		window.wm_iconbitmap(tempFile)
		# Delete the tempfile
		os.remove(tempFile)


#************************************ Program Start ************************************#
# Create objects
Base64IconGen(root) # Create icon
swap = SwapValues() # Generate settings from dictionaries
filePane = FileSelection() # Open file selection dialog

# Run GUI
root.mainloop()